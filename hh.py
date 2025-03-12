import os
import cv2
import pandas as pd
from fast_alpr import ALPR
from openpyxl import load_workbook
import tkinter as tk
import threading
from collections import deque
from datetime import datetime

# تحقق من وجود الفيديو
video_path = "video_2025-01-18_20-27-35.mp4"
if not os.path.exists(video_path):
    print(f"Error: Video file not found at {video_path}")
    exit()

# تحقق من وجود ملفات Excel
excel_path_1 = "detected_numbers.xlsx"
excel_path_2 = "car_info.xlsx"
if not os.path.exists(excel_path_1) or not os.path.exists(excel_path_2):
    print("Error: Excel files not found")
    exit()




def check_third_excel(plate_number, excel_path_3="blocked_plates.xlsx"):
    if not os.path.exists(excel_path_3):
        print(f"Error: ملف Excel الثالث غير موجود في {excel_path_3}")
        return False

    wb_3 = load_workbook(excel_path_3)
    ws_3 = wb_3.active

    normalized_plate = str(plate_number).strip().lower()
    print(f"البحث عن الرقم: {normalized_plate} في {excel_path_3}")

    for row in ws_3.iter_rows(min_row=1, max_col=1):
        cell_value = row[0].value
        if cell_value is not None:
            cell_normalized = str(cell_value).strip().lower()
            print(f"مقارنة مع: {cell_normalized}")
            if cell_normalized == normalized_plate:
                print("تم العثور على الرقم في ملف Excel الثالث.")
                return True
    return False

# تهيئة ملفات Excel
wb_1 = load_workbook(excel_path_1)
ws_1 = wb_1.active
wb_2 = load_workbook(excel_path_2)
ws_2 = wb_2.active

# تهيئة نظام التعرف على اللوحات
alpr = ALPR(
    detector_model="yolo-v9-t-384-license-plate-end2end",
    ocr_model="global-plates-mobile-vit-v2-model",
)

# تهيئة الفيديو
cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
    print("Error: Could not open video.")
    exit()

# الحصول على مواصفات الفيديو
fps = cap.get(cv2.CAP_PROP_FPS)
width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
frame_size = (width, height)

# إعدادات التسجيل
pre_detection_seconds = 5  # الثواني المراد تسجيلها قبل الكشف
post_detection_seconds = 2  # الثواني المراد تسجيلها بعد الكشف
buffer_size = int(fps * pre_detection_seconds)
buffer = deque(maxlen=buffer_size)

# متغيرات التحكم بالتسجيل
pending_recording = False
recording = False
recording_plate = None
post_frames_needed = 0
clip_frames = []


# دالة لحفظ المقاطع
def save_clip(frames, fps, frame_size, plate_number):
    if not frames:
        return
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"clip_{plate_number}_{timestamp}.mp4" if plate_number else f"clip_{timestamp}.mp4"
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter(filename, fourcc, fps, frame_size)
    for frame in frames:
        out.write(frame)
    out.release()
    print(f"تم حفظ المقطع: {filename}")


# دالة التحقق من وجود الرقم
def is_plate_number_exist(plate_number):
    for row in ws_1.iter_rows(min_row=2, max_col=1):
        if row[0].value == plate_number:
            return True
    return False


# دالة حفظ الرقم الجديد
def save_plate_number(plate_number):
    if not is_plate_number_exist(plate_number):
        ws_1.append([plate_number])
        wb_1.save(excel_path_1)
        print(f"تم حفظ الرقم: {plate_number}")
        return True
    else:
        print(f"الرقم موجود مسبقًا: {plate_number}")
        return False


# دالة للحصول على معلومات المركبة من ملف Excel الثاني
def get_car_info(plate_number):
    for row in ws_2.iter_rows(min_row=2, max_col=ws_2.max_column):
        if row[0].value == plate_number:
            return [cell.value for cell in row[1:]]  # إرجاع جميع الأعمدة بعد الرقم
    return None




# تعريف المتغير العام
info_window = None

def update_car_info_window(info, plate_number):
    global info_window
    if info_window is None or not info_window.winfo_exists():
        info_window = tk.Toplevel()
        info_window.title("information of car")
        info_window.geometry("600x300")  # زيادة العرض ليستوعب الشريط الجانبي
    else:
        # تنظيف النافذة من الودجات السابقة
        for widget in info_window.winfo_children():
            widget.destroy()

    # إنشاء إطار رئيسي لتقسيم النافذة إلى قسمين: معلومات المركبة والشريط الجانبي
    main_frame = tk.Frame(info_window)
    main_frame.pack(fill=tk.BOTH, expand=True)

    # الإطار الأيسر لعرض معلومات المركبة
    info_frame = tk.Frame(main_frame)
    info_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)

    if info:
        for idx, data in enumerate(info, start=1):
            tk.Label(info_frame, text=f"Information {idx}: {data}", font=("Arial", 12)).pack(anchor="w", pady=2)
    else:
        tk.Label(info_frame, text="The number is not registered", font=("Arial", 12)).pack(anchor="w", pady=2)

    # الإطار الأيمن كشريط جانبي
    # اللون الافتراضي هو الأخضر
    sidebar_color = "green"
    # التحقق من ملف Excel الثالث لتغيير اللون إلى الأحمر في حال وجود الرقم
    if check_third_excel(plate_number):
        sidebar_color = "red"  # تغيير اللون إلى الأحمر عند تحقق الشرط
        sidebar_frame = tk.Frame(main_frame, width=100, bg=sidebar_color)
        sidebar_frame.pack(side=tk.RIGHT, fill=tk.Y)
        sidebar_frame.pack_propagate(False)
        tk.Label(sidebar_frame, text="warning", bg=sidebar_color, fg="white", font=("Arial", 14)).pack(pady=20)
        print("reeeeeed")
    else:
        # في حالة عدم تحقق الشرط، يتم إنشاء الشريط الجانبي بنفس اللون الافتراضي
        sidebar_frame = tk.Frame(main_frame, width=100, bg=sidebar_color)
        sidebar_frame.pack(side=tk.RIGHT, fill=tk.Y)
        sidebar_frame.pack_propagate(False)  # لمنع تغيير الحجم بناءً على المحتوى
        tk.Label(sidebar_frame, text="safety", bg=sidebar_color, fg="white", font=("Arial", 14)).pack(pady=20)
    # يمكن إضافة محتويات إضافية داخل الشريط الجانبي إذا لزم الأمر


def process_video():
    global pending_recording, recording, recording_plate, post_frames_needed, clip_frames

    while True:
        ret, frame = cap.read()
        if not ret:
            cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            continue

        # معالجة الإطار ورسم التحديدات
        annotated_frame = alpr.draw_predictions(frame)
        buffer.append(annotated_frame)

        # الكشف عن اللوحات
        detected_plate = None
        results = alpr.predict(frame)
        for res in results:
            if hasattr(res, "ocr"):
                detected_plate = res.ocr.text
                break

        if detected_plate:
            # حفظ الرقم إذا لم يكن موجودًا مسبقًا
            is_new = save_plate_number(detected_plate)
            if is_new and not recording and not pending_recording:
                pending_recording = True
                recording_plate = detected_plate
                clip_frames = list(buffer)
                post_frames_needed = int(fps * post_detection_seconds)

            # الحصول على معلومات المركبة وتحديث الواجهة
            car_info = get_car_info(detected_plate)
            update_car_info_window(car_info, detected_plate)

        # بدء التسجيل إذا كان مطلوبًا
        if pending_recording:
            recording = True
            pending_recording = False

        # إدارة عملية التسجيل
        if recording:
            clip_frames.append(annotated_frame)
            post_frames_needed -= 1
            if post_frames_needed <= 0:
                save_clip(clip_frames, fps, frame_size, recording_plate)
                recording = False
                recording_plate = None
                clip_frames = []

        # عرض الفيديو
        cv2.imshow("نظام التعرف على اللوحات", annotated_frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()


# تشغيل واجهة المستخدم
def run_gui():
    root = tk.Tk()
    root.withdraw()
    root.mainloop()


# تشغيل الخيوط
threading.Thread(target=run_gui, daemon=True).start()
process_video()








